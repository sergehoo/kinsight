"""Export du tableau de bord Gouvernance (Indice Groupe) en XLSX et PDF.

Rend un rapport (cf. `build_group_report`) déjà filtré par périmètre RBAC. Les valeurs
absentes (gouverné, ADR-0007) sont rendues « N/D » — jamais 0.
"""

from __future__ import annotations

import io
from typing import Optional


def _fmt(v: Optional[float]) -> str:
    return "N/D" if v is None else f"{v:.1f}"


def _num(v: Optional[float]):
    """Cellule XLSX : vrai nombre si disponible (triable/aligné à droite), sinon « N/D »."""
    return "N/D" if v is None else round(v, 1)


def _scope_label(report: dict) -> str:
    scope = report.get("scope")
    return "Groupe (toutes filiales)" if scope == "GROUP" else ", ".join(scope or [])


def scores_workbook(report: dict) -> bytes:
    """Classeur XLSX : synthèse + domaines + filiales + tendance."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Indice Groupe"
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="16191A")
    title = Font(bold=True, size=14)

    ws["A1"] = "K-Insight — Indice de Gouvernance Groupe"
    ws["A1"].font = title
    p = report.get("period", {})
    ws["A2"] = f"Période : {p.get('start', '?')} → {p.get('end', '?')}    Périmètre : {_scope_label(report)}"
    ws["A3"] = f"Indice global : {_fmt(report.get('global'))} / 100"
    ws["A3"].font = Font(bold=True, size=12)

    row = 5
    ws.cell(row, 1, "Domaine").font = head
    ws.cell(row, 2, "Poids %").font = head
    ws.cell(row, 3, "Score /100").font = head
    for col in (1, 2, 3):
        ws.cell(row, col).fill = fill
    for d in report.get("domains", []):
        row += 1
        ws.cell(row, 1, d["label"])
        ws.cell(row, 2, d["weight"])
        ws.cell(row, 3, _num(d["score"]))

    row += 2
    ws.cell(row, 1, "Par filiale").font = Font(bold=True)
    row += 1
    ws.cell(row, 1, "Filiale").font = head
    ws.cell(row, 2, "Indice /100").font = head
    for col in (1, 2):
        ws.cell(row, col).fill = fill
    for s in report.get("by_subsidiary", []):
        row += 1
        ws.cell(row, 1, s["code"])
        ws.cell(row, 2, _num(s["score"]))

    row += 2
    ws.cell(row, 1, "Évolution (12 mois)").font = Font(bold=True)
    row += 1
    ws.cell(row, 1, "Mois").font = head
    ws.cell(row, 2, "Indice /100").font = head
    for s in report.get("trend", []):
        row += 1
        ws.cell(row, 1, s["month"])
        ws.cell(row, 2, _num(s["score"]))

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws["A1"].alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def scores_pdf(report: dict) -> bytes:
    """Document PDF A4 : synthèse + domaines + filiales."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="K-Insight — Indice de Gouvernance Groupe")
    styles = getSampleStyleSheet()
    p = report.get("period", {})
    flow = [
        Paragraph("K-Insight — Indice de Gouvernance Groupe", styles["Title"]),
        Paragraph(
            f"Période : {p.get('start', '?')} → {p.get('end', '?')} &nbsp;·&nbsp; "
            f"Périmètre : {_scope_label(report)}",
            styles["Normal"],
        ),
        Spacer(1, 4 * mm),
        Paragraph(f"<b>Indice global : {_fmt(report.get('global'))} / 100</b>", styles["Heading2"]),
        Spacer(1, 3 * mm),
    ]

    dom_rows = [["Domaine", "Poids %", "Score /100"]]
    for d in report.get("domains", []):
        dom_rows.append([d["label"], str(d["weight"]), _fmt(d["score"])])
    sub_rows = [["Filiale", "Indice /100"]]
    for s in report.get("by_subsidiary", []):
        sub_rows.append([s["code"], _fmt(s["score"])])

    def styled(data, widths):
        t = Table(data, colWidths=widths)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16191A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C7CCC9")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F2")]),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ]
            )
        )
        return t

    flow += [Paragraph("Scores par domaine", styles["Heading3"]), styled(dom_rows, [90 * mm, 25 * mm, 30 * mm]), Spacer(1, 5 * mm)]
    if len(sub_rows) > 1:
        flow += [Paragraph("Par filiale", styles["Heading3"]), styled(sub_rows, [60 * mm, 35 * mm])]
    flow += [
        Spacer(1, 6 * mm),
        Paragraph(
            "Données issues du Data Warehouse (lecture seule). « N/D » = source non encore branchée "
            "(aucune donnée inventée). Pondérations : proposition à valider par la direction.",
            styles["Italic"],
        ),
    ]
    doc.build(flow)
    return buf.getvalue()
